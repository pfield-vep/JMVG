"""
pages/3_Data_Export.py
JM Valley Group — Weekly Store Data Export
"""

import streamlit as st
import pandas as pd
import io
from datetime import date, timedelta

st.set_page_config(
    page_title="Data Export | JM Valley Group",
    page_icon="📥",
    layout="wide",
    initial_sidebar_state="auto",
)

# ── Brand constants ───────────────────────────────────────────────────────────
BLUE    = "#134A7C"
RED     = "#EE3227"
WHITE   = "#FFFFFF"
LIGHT   = "#F5F6F8"
BORDER  = "#E0E3E8"
TEXT    = "#1a1a2e"
GOLD    = "#D4AF37"

STORE_NAMES = {
    '20156':'North Hollywood','20218':'Mission Hills','20267':'Balboa',
    '20294':'Toluca','20026':'Tampa','20311':'Porter Ranch',
    '20352':'San Fernando','20363':'Warner Center','20273':'Big Bear',
    '20366':'Burbank North','20011':'Westlake','20255':'Arboles',
    '20048':'Janss','20245':'Wendy','20381':'Sylmar',
    '20116':'Encino','20388':'Lake Arrowhead','20075':'Isla Vista',
    '20335':'Goleta','20360':'Santa Barbara','20424':'Studio City',
    '20177':'SD1','20171':'SD2','20091':'SD3',
    '20071':'SD4','20300':'SD5','20292':'SD6',
    '20291':'SD7','20013':'Buellton',
}

# ── Logo (same AVIF as other pages) ──────────────────────────────────────────
_LOGO = "data:image/avif;base64,AAAAIGZ0eXBhdmlmAAAAAGF2aWZtaWYxbWlhZk1BMUIAAAGNbWV0YQAAAAAAAAAoaGRscgAAAAAAAAAAcGljdAAAAAAAAAAAAAAAAGxpYmF2aWYAAAAADnBpdG0AAAAAAAEAAAAsaWxvYwAAAABEAAACAAEAAAABAAAM9gAAFvcAAgAAAAEAAAG1AAALQQAAAEJpaW5mAAAAAAACAAAAGmluZmUCAAAAAAEAAGF2MDFDb2xvcgAAAAAaaW5mZQIAAAAAAgAAYXYwMUFscGhhAAAAABppcmVmAAAAAAAAAA5hdXhsAAIAAQABAAAAw2lwcnAAAACdaXBjbwAAABRpc3BlAAAAAAAAAPQAAAB1AAAAEHBpeGkAAAAAAwgICAAAAAxhdjFDgQAMAAAAABNjb2xybmNseAACAAIAAoAAAAAOcGl4aQAAAAABCAAAAAxhdjFDgQAcAAAAADhhdXhDAAAAAHVybjptcGVnOm1wZWdCOmNpY3A6c3lzdGVtczphdXhpbGlhcnk6YWxwaGEAAAAAHmlwbWEAAAAAAAAAAgABBAECgwQAAgQBBYYHAAAiQG1kYXQSAAoKAAAAA7efQ2vkqDKwFhAAjYA44kEg2AF6CPeYU45VXAJ1Y+L8ltPhLa9Taur8JYUPf+O3RJoTgnB7ZXhbEt5aEBF36NyWnN618Ir6GA0D2Mr3lhPqmlXykbY90y6VdiYc5sAni7WHDhYifcZ3O9XZqd7FaeU8jo+Ec5u2+b/CZfA+QwlxbfmpuX3cddf7qoXkabfqdbqTi9bapSo2Y4mnZK/IcvcH3hg8JbsYwc0ysyikhmsDSKvcFXdMiKjlg36ICOb70hSVdp+pdviLjHBzvJGG4hNfwEyuLb5k2sLxmIbzdI/Q02BC7X1204z4YCG8Acj5+GE8GShyGcfl1ziiDTXRosoGV4ccAM/N1nLWL4bNWmdgQzPd1DNjJpXW2OGHjf9dJYE4fY34AkI1CHCc8fyBgPvBpRRWIk1KORCYxWytPhvOilAH2rJEo7+4E0A2J0NV/gil/qxF2DoyrjWn3n/A/7G4rkWNDIspVE4EXrvOsmMr5Ma9jS170Q7bC/uSCGbBSwsN+i6vi82lFCmObYsO4hUEG5DJqSpLJqobSL/mMmSQxU/NjcM4tEGmSOK7gEZ1UVC9f7ZLiUmwBt0i9WuMOlqCqkOLFvB2pFKlOM0jR1z1Kv+wbu+MojMAkbuNTdfoRZ3Q59QltdiQkYVZq/2wUP6g6bXe0xfR4pX1Riv5h9fvlKsGHrVbhrXRcFimfo4i56OA8O/sYOfCqLDUbhuz47UJut3SFimHAu4BKufjbKyeypwsCLZ6k/fo1kl2psGx9ODPvZTS+PMPpvKpGZ9f408PbI4VY9xA7X1k3jqLqkEjKioCZY7WKqg7uY1C9jMQjocRvBKK9KrpqxejFxYh4LT3q/05w6sxIC+IzGh9BrgSt95Of0BSvJRD3lIJE14kWI4fC/kcAHJqzBII732LDO1zNJTg9Ome7CAFK1Bea88qRuqpaM9ReC7047B+iNkKiZkZkGjCcnaeiueYz/IMnjpQuRj1t+XZkZLjJZ9P4Iti6ENnclhVc5IGfgYRPBZ15Rq6Og3Nnfu+kAsgdc38L6nE90RiBhlRcXmrwoTgeh/VWNDL8boiB0OKjiZxZkNbn+SxfIwexVu9JmBnW/wTSrx0BS2RuLh6fVG5neOKsAhfFSCw+vYJrI3V3Kz17kIcaH7ZOFcjBx4yI5NFtEGaAJAxfT6uC1Yztv/qMPVie/mS73LKhbWrYmQGKs9X0Gnp5ZS6JtoSoGF3JxZw2b56SXGV+GAM3gTBsvuqIHaf0vDupzFvyCIboadKImZTZGiS4ugKzoetOMR7Q1UUDJub87JvPZElrVZCwEVee256GQv11vTQ+WVlXzdRQDFExn31JrBfQfIEaMow1vtjz5fbAp9v/YjKzUzqnMHMXCdPU6fz7BimdPwDEg44fKBkcxxtOKD+j8IG0O1rn+Tbj7uwFzg50o88RcEoAfg/HkjrMluYPiJXjYfG3ag3F47rPPyJcCxiAReoSkiadGtURB63q/N+s3bW0GfEKD5Db8xK8nBGAmm/XFarAXp9Z3MDdIx0TYtez1zLX4HKidLwijUMMJ+M7UfzzpIyhoCr5LVhAerRddXv37PqIr8fKXLQzZ4xNjkNyYMUMr7UwKcRD5EDe9VdQjcOYM9dlOhYQVJClZ4zQ3+64OPyTqwOK2Qbz8iaOkRCp3kBmKiRGAtcrlreO9vvb6ZONO+v1AcOz+MFc0Q207bmwlpPRjKMr03LnlOZBmxe8bgSxo+qusNAu4s+T0TSciNi5EcYLt0nmEV9141PS/6TxVeg6D69XdwhcvRtnKVS/OCnUofSbRXav+XeM7CKjjOsOw4ZzO+DkMIyGDzdOOP+6BhdaZawMZIwtbjcZ4vgZmcoG5eBc/kUzgJZtw9pcC5Lh1z+grFBk3rqffr+kdSHGx5lX2krL0ZXCTHAXRBEVeCJN1j/zZVvAqhKXRwKDL7mK2ZVj/He79IBS0wLPOd+qeY1hB0VxJuLUtWz7kr0yR5lYCAN65MCINH1vrIkZtMos8QnI9km6kX7n85eYy5v0Q7NZPzSfwH5SGprqmAmGxNljZZK2y2tSUdrGaYZV6BMoSzz/imeMfJhjyhtAfFhkeb8EX2tu/VdH1q6muB/3dj9SSMRL9atkqfeSffylBhCToDeMtLRhxgklRsWcdJm0iEiapuFJRFAm3rjYg1clFupgfcHwj1HLqPi8a34atBZvVoGp2T6TNTGM/uOodhWxMYxRujE1Nj3+9PzXG9xiEoch+PLv7QVvQqceGUXJSgCextASaMlVmVROkpkDsHjCz8w7XW+YZEkcdd77S9EmhmG1GxwT2MTDkiQboHFZvmi7eyhaLKzL0vhii1lAIpRCSYFCs31FaS4F9001c3HcxFnPpzzDJZvmLd3PCtqR5dmqjjdPCY6SfZMXL0qj6dKhPq5wS6K93gMVtUQcErIo7QqP8twQb9FnzX/mtwfO9/gW73HFDzEh9giS5CgCRgM8y1YHdUOA2zdbAuuzvoTeAoPq31faDGOAKxhhqoF7XC0BIGCikyWDTpV7sUv7rMjhrAyV39aqvKhLGBDlbL9gTlGHOtQzQWutC0xkru8wv23FS+/X901BHsL8tjxLAp7hqaWF6f7ZknyuP8/gEXf53wrpB62iaG+ApTTpdjvhg6nLwO4LmwVEkEQ2qHRkqYUvKRs8eVpjVslNrrDeHRMj2durWybaKYrnRNBZMv/e+cDih2/uND41nGozotfHQg8BVOEWgHRAVM97WywzBGCnz8a8gYlpbi8JA1I9ZfXgJxRQ0hgkmBalQhOFDBroWccUDtL1LUTT+LhNGnnqUGzBi/vp+veAqOdlm7HrobOImzbLj7tCLUc1yQrQE9TWI4NdhXtBqhpJFHjZP+Pi6FIFeUStkLCaEu+rha7rNfXFaAKqNT0j/fP2KE4E4toj+sjG/VrfrNcfgxi46dOUKA2SieNai3bjopH/DsV+rQnH/P/SVMhRm6qBE0tUzbaex+BWjrzVYa7nBwi9wPsTBfZHQF8zq6FfuDrLqz0uq/GfTugt1O/vipbfn+FBAdAF+czrlbmGL8ELvGaMjm2fPnb0RsjPtuRIxxzMT++6ynwiYUL1ciCC/6Znnj3gYfK/NkaO8ngX/cqvsb8xVcniYpj6Dg6k9W91zVakxk+0Rw2twM9KqYgjumDhFRlFTsvV3/3WPKpf9ntP+KG+r+9VlgLBKznzI5yZ9I3z+7CMynULd+67FYIKBngnjKLZYVbeIsDN8japLuQlZHMscDK/SIqFWMvz1p9GYHKen8z/7Wrsxz9vAhbBgEjnmxrnmafrydVCKIJzW2U7R8xbs4I0VcLTt5UEsM2ETvWs/RwZ8WLAcmHij1yQgBwpww6w+ydeB84RiqK+eHLfXdDwvYOWXzVhGdLD8wrLahla0qgW0IoRzpu1yNXW/eXF884CgRhJjFnsFian8oAaCFVw9x8czi0DKrJxSFd7tYN2REOm4aieQ35J9EgiFPXwZnJzyiQ118ES1vVtXYVjAmxwgJEQnHfhIUkG/QcblYAmo6mxwvrd1ZpQsPQpoSIypY2B4QHYOmsWvRLe7RQCArC1Mejlh/CXRchqgrQXknxDFgf7FmZKfpMvATa0Dr2mm5xZkwpazrnQ6Ye0tdIbMibURUwr5yZZ1qi1G0DTLPur3K0pBUOE1OfnU33ehUE3BIcad7Sf5td5+3RtXtsWoZp6bzFYht9xbIAjYSJEls8mPC+yRD3VEzQQUSiiUUEFBBQQTnJZXE067t/tZNvlh/L/0H1PO7MilKwgVNMG3X+skYXzc0Gk9Nazfxa5sy6GFSkWKAkCSeNRl+IZFSAEgAKCgAAAAO3n0Nr5CEy5i0QAIgABhhhiQQS2Yo99QQZ3yS3E4QLPT4QkpYHg1AqFjnlOsAbuQSga1O4axF+fZEUBRt72l86Rz4GW0ijXtQNWKJV8HnV8iSTb7fZBy85EOsjrocBIXoklKe89dvN5eOflnr8mo5R7MUYlMKyPyChxyaNGFS1kSjbJkhiN1UWbYRbvwc0kfvyZogT8yRzvcgTlp38iM0ZoPE/upSC1G1e9pCC9I/xDNPftMb8SZeESNsAzSrFFt/hfX4DchcXGZMKq2RJVrMziGiGfkyaAlYMql+eyFsid4FjaB3gCTvuAUdYHDrkAo94h91dhaxSpe3TO5g9mSWug8dvchFEOyXrgI63qFG9RXc3O0UcAKh76nn6D2SBa1Mf8XMy2oA+7vrPu97Ek8wMCl2ijQlbvKTZFMakcaV77dYg5cdQX8aPvL7ouhb9nBf7/JGM5zE8zfKtLt0JsOjHyUN3w0HZO8LBYKCtXfaQ9vfJsj5MZUErE18z9mAGWSk99GozJTL44Uz6swNadkhi3PSYU7UkCmd9qgp+j4HcmWY3IMzQIkwpMbMVJ4uK7NRlJL2Ei5gLAQE3XpCAFwLUx00H0iVdqDxIW7hbErC8AxcJTITZi2bsnkldQYtY1NFmXthte6xYOc6taIOkWqTPDjol8huwXf5UxQeb/oPtgEq0f8eH9MhF3lWFQuKJzoSp6fSCkbFIhq6T36MqYWtGuAvXr5CAAPz9XfkowALEyO6UdFPKKUaBThbJvB5PC/D3vRk1IWZst0s04KQ8CRkYcjwZPeDlT7klbh+YDdpUixTzfcXcc3j+/Zf3y7/osepZCs2gqOcwstXNtrPDdxbbGX00knIrdEscUj3RYk7jJyLeEWJ6KGW1FvUIsOQxGpjMtUP6sAtEwQnarn2+B6IfjbC/QDfNnSXRCzkVZsakcidIx99ZPpMdTzh1pYlHohXE436fM2YMXSW2heVidXLSpWHlVsf8uOIU1zmnLhjR84lIJWWXr7VAU1iqN+/NiqEkHq0P9PD90w8WJCCv7m2gt1O7BOYuZK6eaifyaX+6PhFkfZFhJvrEjf+Zv6Ge7yU2c2nmjcJ3fB9FLo8rLWXoeNF1XYl6jcEDrkYbrH/YgmazhtVq7XGDTEZ/KZVhcxPTzJkySuPCw//8FSyCQqIZ1haJuK1Mix9OPrq2XgYFTVt+awpBEtnwiPb+3ikannHQi0lMjvPhCoLAAjwl/VFAvt1wiuDgYgUcu3ns5Z69J5yXiNFrXym6L8/ocW/BuQk859LbjKZxTGAt2gcuxabdegZEcgFoSQQHyjpkApxEA39IoUrhcEtVlkj7QAB3s3Pbw88JpLcKtRjcLMi7O/vGfGyuI5N6PU6z1EoPacmCv2coZoijf1+d67/3oF+vpZFKtGXGRC/dnfi8xYvfkfUwWXUqtxEASnRa0gR5yhUlvszRFsR21UXteaD+ERGZamyKUdlcu5DhokaIIV41J9vLvqOcNBR5akDKCjvidG94OTgdOCpSaNANTPvjP7h4fZaNnQ/BCwOidykKl64WP6S1L5+P8TBySOYNCP/mg5YZp0SJ3Cmt73CBKlpoGXuYfLb10vQgL5oErrkzqY1Nb9QVWbptKM3JeZP1V6NhtZIe5BnpvmDiEiVkGg/XCUfIQhWx0Cc0qAuimG0rXgSfGSCA4cYlIYFTMKRCxkQ3eanXS4gekcRxBb/NO8pA98bvP1VshLR71N/3DNH96akSOriqr14vR8HgMOvbUE3485tQNhJhqYi2CCOjeSbJ5VGXJOBfyoI5CpCcYP6jHW4nZ5oVNZ/q5e+UPCPjGJODUFDekyFwu6qKn69aNuGJr1sAJoQIArxtg6C8r4ABKEZt4lMRRhb0zMcjLv7lbhaVy1R8Y3tJqsa2xjQuyqFAptUcmBN+IiIpzHjfXzO+Bca6B7qAlnzwvxgjEbQ8jNT5BmnztX10ajlGbKKAYyu46B/QoHibqipWfv7WVi5gN+Z9cjQsIZp59gu9c/baomCgHB/suyz1izOM2j91mcxdvD9iwwx7RHaO3I57UKi1cplndW5FHZLTjyyEGsJzvcTZTg16SYBfekFZ/CGMQkb5BPCa+GEfByTjyC4x/f/nAYQxprCeFoRjWYEb3dqWVEZHQQpEjquTY+3r7zfjeThpteidRVKMbXtWc+4EsEuFpibqXcpjeuptLPF62yRjsQcvHyM4iahYkSLCgqLpT5HpEGIPwUx/8IB07QTlVgLEL+haSQdSuwNtAkl0LI/8ZMUFHmR4eE2LTOJjEi+fJb37KVHenQDBz15CHj3Ar5juFL+u+VmSLU7QaCdlbe5+PqglL277stDi2ok90GNoZ3WCpQ/UfJSAFj6HZSxYFbRzPoVV3lYONtvsTRjyfMSPTOemzovyCIlVGnTnVo0X75y4LEQPwdyKF49Lr3jLdWQTRYfd+SeHztWKVYY4evsc/BuSpTzdLq4mc8bp4VkEIQZTwPnPludBk8F8jLWvHFkSfhD3GWcts4/IPuFTtWtL4EhSVKLnOxrLYZiplKsAZwXYs3DOzdYpKRFt9lkjeFIWdvcsZ0kdL1OUSvoOlG0IxBcrF6NnESOEHrpAXhVpkXp0m5IX3ukS0/2lpCAoIHSzp9nUAAAACYpnT/s4JvyoMemzLwK1XsPWy+SxknPQAX3zBXve6ATk8Hixz+SjqBvg8D0spP5N4sjVvf6o0VGTqLbjFbrzSJLXIdq+cuIKx+lUH4G87zSMU7TdQClB1TXKYScRs75YaCUHpCGl3TY4dLYRczCSO7HhybScdMxmmjYa/fO37e98C2dfiinOMKfet0qSmCnwJkKSVobQpz8ubq5zDN7UVw+HPZIuhRHVJXVUuM5IEgo2cu0eXh4BB2fYPBoGzcx9mtUyClQILpFDjQSxfSECq+YBK6QzG+3Ao+TBcOyhIjdiIXIDwwB6DeubdQKqmpH5OoAgwdUy49CAw8QzUi25jdhu/Gze+fvdWR2iemsUABOIlZ4Z+SVQlYh5mC3yzc4AtjNwpbDUFeQJTlQ5uLKGycrkLzVLeMQ1rGsdpOJgV7ntBsaf9L7z4WJ4InGZs/IyJcn+O7PKYqPcXVKHn3V+JjbXs7Vyu5MOxJgUDvvbzu/4CMbSiA9PRxHKWZVTiSnhDaRR5ABmxCIDFZHZgusJTOegGPc7Q9nqb2siH8HBIWAFHX7MuiWln+1NBBNzUOiqRle1zpYz0EA0Mv+MGH/K/2c++Bf/Ct7IUfERgMf/TiQr8JKaigLdDwMXTblA7jQFSLT+ncEN924xSryq/Kaj8wbsLyBguyi+dUCtvCi7HiSfgjt43f59ve3puGqFGOpfi9SIwU3ocFZw+IJhIs0ZJFDeYdyypnuwW2mwIf/ArjmObBRsB/4ZdN8/VCU6NCJajrLSudwWAPy+Ac7hFl5H9nb42bClEzBMG+4MtySYl0lvz4aY2vdlImlir1vOaYo3mU0QP/G9ij4XffYVrKsPlEuxzh5U23+9XrZF1k8C/bm+pqVxlsk7LOzYwpxE8ePITU5URG7yLDWKLgGX5jkNe2oaFdz+jzfOvpWP73E0YRciTY8zHHfmwupzkLh5KyQBRz7SUIfKtIYCBGR121uodIPxFVh04HafpJ1RM9dU1y9CPI5SdIhmYScx482OiUKz//WUJXN4/Hk6fmXDDMDQK5iG0IrgsdLt+EsJq4EcyCirgWUD6HNX9WGVP2T3i0Q4g65/wRH1/SR7ZquXrTG3UC+dZY9vlKo/uj+wXs7MzugqZFVb5yJ/HYC3EvCkjs8ExbxssuqHHpvFcdlU1/8ZeAicfepZggMSXKWRSs1pdqPzo6+AY5lB/gdGnROmL4VORqrw9vbjoWAkpUZHZ0MleT2Zzl0Rupn0TER9eSsY+/1nI6/R0Xvz2/Bc2XYyXqIkywtlVaHgyuvLSGsvrAKW3uqiATqpt5CtX3WMyE3i5EmzLuDwX25OqmlCBYty4+IgEqk/Ot81pt3SU1sNjux/aaXb9agef5fKyko+x0pcAa7x/6NXlw4M3TzlHfFTrZMQlyn4IGGTpy1F2T8AYj5e3bHQeXJwDhVkiipqikQDFTcR6CDRfjrUGI7ETtbtEMqMik8NYUDtzVb0okfKAnqmKvclUwOOFbGd38tlzW3pjVkUDWruBQweCpVfWoQTq6alNHSm6FUBcigi6EPFKFAyZgpN0AekoiG38fHulncfnWZ8pPX+Sj9f0UlGTvP0foXyZsQ5U8K0PU7FdKDI4jEZZZjgo+M5LElsB8miyEJ1KB3yA50gc66R8L89SrMwlMtbitSnLibIE3bub/eOymVsFtnVMn/Sych8lCaAwXUQioz8hjpYtkXu9q+D2QiTp91ofDlIPhfM0vHpDOXdka7S/YuQnRxyxEfcZIO6b9HCIk8MzwSYck92Ut+Lr9hANJzirt39owXDt3X+xWjplbX3ieEqPDq8ErUaO9etMcnKSAOiNlfls0GZP+zGCgcTzjqbyMYSxftkHKSOBCvB7D8zSyBt0tSG+DC/LsrceQ1DvkkKwJNJbRXmbj9lDgHz294tMKCPq2OZHH1ddVMOCSlHDGOSn61bjKA/MwgywiIJu4wKQD9p+5rx/tsr7uJg20MYE22whDk/HG7ZRIr/7Jqq5KffQ7iyKDvUXzCNCO3KDXjBGmYjge7x9Rp6fqlhyrnqyCeu4ryIVxeXj6cwCV5COkCkGA5PMdwNYeEUKZ6TM9HJsCPXP91W3ohJ96zFNJIMbVm1zofN4U+2m9hNzCtH2Gnd1NGCLG8ZoDi6LjgjUKy4PYESC9y+VhhvEtYm5nUMHuakTC64D1vo4kbZuriLXnZs//GIvoqUkGFk5JpxflnnxRiVOhnQolqjelrSGwvfHhQxM8QBBG1EgYLhM1vFYVnDBoXk4KMWHrdrsRmyb69IY6NfL857TY3qa7+bOLdlxbteSJHe72SWX1GSKn7m0PzwZ9kC9oKUAnMmdfnHrStu/NFEyLnUMpKrfgEafSkiV1s7g/UK1WkE2Avh21x9CEzrJf12ozu7UzPhIbGLVM4o8J8+jCV/XlqwJvPvagD3Ql6XSC2ygJPmnWb3kUJRLuFaiU93GDmkb03C1/mCJQoiHL6C0LdlHVdINbIjn3G98ib075zo2rFMsVEIdxg6FUUE0mBtJb6pgSKp3tXVlCEO4abswn9suIup9WkG3Lthciln7oR5iwDBunAG8Edky+wbvB07MjiAXwx+oTV+ojkgAGoXch0xswqVTRqDjzdVbdLbqXt3/LtvGjT3szorl/GJSiXuQ/UL77FqAs/5KEqMY230Reic5CcEzmORJ55ESZgDtO268NguVJBMOwQx0YkX+aHBC7rQ1zJ29w+XkaBN2zfNEbfGwxTMA1/tpIQfyjiNZMvYjVa5swxmdUTqYJ5xzAtT89Yxrl0o/wVE0C6xuJYDc5rp5equDYKimqARBVG6e+Hc5A4IcLlI+uKOJr0yVkGog5N98Ic8hcXaAqeND7H+oUtmricfvYxs0vLk8ZAOjNrldbAfWyc7QkvTNZVf3/RkFQf/v9QEyn9b2Lb35sCi9diELqkZgDdrt84UUaz9lhpVXqkQpqLkHzPd1Qwve2wTsAxA7zZqgeDaexknNuOTFlQyhXaLfi6YsUgnLekXCkkg9iSgVwuZWpC8r2IQcljn/eRo4clZcjP1vZipu692D8i7PDRXiYTM5Zbzf0H1F1RTHrYkSCqzo1Vz49c7C5mzPdFbLUUwaJGj2RVBqWmKgYneEqa6q42VwjJYCMay4FfOkk54mxOyTxEs25VnzWsvDwUaUkvJtlywHpdApBlx/pWou+ur5gaho5VyeOlgq4MU13nuF1p7aRNSDEUSeXve4x7FOqVm/FMzkVbqqVx/GmxpOQOYmO0kNG+buRm8wB6NCncVrAW1x1Cqjz0RgT4NZCvKeXF10FEqdrqngJdPubDfuR19wbsc+9DOPEeMTyUTkT1Vgu9bD2rvH1qwxCxqmmKvm0w60CIU5l3Y/e4qKOUR2q+mB9y/L8c7h4BzP9THdS0wKNSGd3ZneGcyvMcB1p5NT3vwzBau8I7/pugKcFiS0XxSUmrxQY/coluRzxB9gDtNPShuBz+3ouwZz8Wf0ZRZ8/gmcWyhQHxoEHnJr+tysFPFeizXF5AhRJoS/yGeuqzW9fH5hvlDiTWkLOxDobliSTtOuh6ruO6CKuJujAV0CJejLYT4Jtc/AfCqSI2iI2nfQELFhc3iNX4ujmKGweQeCvpTqCBsvf1mNbSG2LG0RdIKEXT3OFAQQRwia2Sal7psUMD+pSE5GkQcDzGhWu9qV//p7TCzWE7YuyNDcphyFflOGosBkoXxCQuAc6/y3WD37Pj3QYgqwypfdyYJaaAQUJZ/M/+fdD/f8GPhI+wUX5TSyiR8kCDu35wo59rdmCNAjanxVRF1bTdEOkjR9z/ujpQJC6zHoMZEsJlav+RWB5U9rb6ZE9YT9FCU2gW0U1JssBDCIki91qJEMH7/zq9S1YZZ7BH75+q6UXEhdr341ZofqN6TtLksBUhcTI1rQwDdcTnxAawwkdC0a0ueH9ZCQVUVbkioqf06PTb12OqF4bA2kaK+or4lMsT0N3pcKRXm11MmIdg6fiIWa6o+EqPwTfdHbJE5L5Q4lnzKX6bTG+irXhlJB+qRqLyP/dZ+yKoEvGi2aDb0vgBKHrxqEJ+jG1rJTEHivPfevoFUAqvgkn5EcYygIheiI0qz0cOp/bIJ1hepw+g98ZEXLjFgyne/tcNhHhwPQvjCQwcHxQaDrUjgfRePsLRu0+Tx9oRAz8xYLlf5It/blocwFHR9mlbWmIVj0fOBmw7J51sXTbYVFGTr4RwTkBW9cxPzSiumKAusLlY+Vw0NivQbUpOFjdhx5TfQdx1dOOewRPTrWU3gYVhmsl78udE04ru5N7BZuztqnTRAkj2Iqv49Qd/vA11TIUAMhdtEWbwtGIOaNuhoG5cHDgCmH9wI46l2Kx9e8N8zcvCfqZfhFzJ3PiPmyPUWOubxacft3u9lcRbvVkvP8L+b3z2btZUB8bxP8K/grX0RHJ90jZ/Uu+MTJYggWvuw9Tk3xTrXIpBNvdDvpblZYaFwYL1tEnH1ZSqom+C/3U8uOGGt8c/Orqp7HvAIgv0fRrPNwWJmfp9Q5Syia13RyS1k8LFOFwhDgviqvdTYfID773979+EPPSUu8ht/xfvlSpKQvx5G0HARr/NB77deAUJFkBtXstXCSG3THa3v8qz1FdcXnIu7XbFFdsC36oc40V3akK0bFtueplRRSxiSYru6reuyeglkhi1nmPYCcqvlFvKhoxZPrPwThb895SasuZ4lu/flhgzgzm0UHsR/f7M7tmh16uAWOrSnyKdHUwUq0wnyd+Ov3K/uoy83mQ6AJtlAHyh4MPYZ+yzrAOcNlcM55ZmxihE6oKoyegUHWkRwO4g8qdDPa2MZJI1OfM3cuxOuVV1/2YJPcgPDdN02wAxtlHc1Vq6iVwUAa9KADxOTpzWAUCBbE7IQ5zsUU1mKsxB9xQ4L1omuOkgLGKg8NobJenXfxA7eQcpjG51v7jYoD7VsVBxn0k426TPTgsvwv2sE/qMhyV4hBT7j9TgJblYWP4LrH8gp23DI6IL+2IP+qTe0PT+tDIFP/ERbn8vqHn3FVpyfrFDHJ7RbxM520EO37Z60IxG7sJ/8UyRGFDsuZIQEZ41nawO/ia9h8AJQcM6MU9+IpP708RHJvP2jflQi4sfPiK8AWfdrqQ7KZLK95CmKJAesgqVyulqC5T/YAhx1O9ktHcoHaXNUzYC2HcRPwA2g5IA=="

# ── DB connection (same pattern as all other pages) ───────────────────────────
import os, sys
_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DB_PATH = os.path.join(_root, "jerseymikes.db")

@st.cache_resource
def get_db_connection():
    try:
        import psycopg2
        s = st.secrets["supabase"]
        conn = psycopg2.connect(
            host=s["host"], port=int(s["port"]),
            dbname=s["dbname"], user=s["user"],
            password=s["password"], sslmode="require"
        )
        return conn, "postgres"
    except Exception:
        import sqlite3
        return sqlite3.connect(_DB_PATH, check_same_thread=False), "sqlite"

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<link href="https://fonts.googleapis.com/icon?family=Material+Icons" rel="stylesheet">
""", unsafe_allow_html=True)

st.markdown(f"""
<style>
    /* Restore Material Icons font for sidebar arrow buttons */
    [data-testid="stExpandSidebarButton"] span,
    [data-testid="stSidebarCollapseButton"] span,
    .material-icons {
        font-family: 'Material Icons' !important;
        font-size: 24px !important;
    }
    html, body, [class*="css"] {{ font-family: Arial, sans-serif !important; }}
    #MainMenu {{ visibility: hidden; }}
    footer {{ visibility: hidden; }}
    header {{ visibility: hidden; }}
    [data-testid="stToolbar"] {{ visibility: visible !important; }}
    [data-testid="stExpandSidebarButton"],
    [data-testid="stExpandSidebarButton"] * {{ visibility: visible !important; }}
    [data-testid="stSidebarCollapseButton"],
    [data-testid="stSidebarCollapseButton"] * {{ visibility: visible !important; }}
    .block-container {{ padding-top: 0 !important; padding-bottom: 1rem; }}

    /* Blue header bar — same style as BSC */
    div[data-testid="stHorizontalBlock"]:first-of-type {{
        background: {BLUE} !important;
        border-radius: 8px;
        padding: 6px 16px !important;
        margin-bottom: 12px;
        align-items: center;
    }}
    div[data-testid="stHorizontalBlock"]:first-of-type * {{
        color: {WHITE} !important;
    }}
    div[data-testid="stHorizontalBlock"]:first-of-type label {{
        color: rgba(255,255,255,0.75) !important;
        font-size: 11px !important;
        font-weight: 600 !important;
        text-transform: uppercase;
    }}
    div[data-testid="stHorizontalBlock"]:first-of-type [data-baseweb="select"] > div,
    div[data-testid="stHorizontalBlock"]:first-of-type input {{
        background: rgba(255,255,255,0.12) !important;
        border: 1px solid rgba(255,255,255,0.3) !important;
        color: {WHITE} !important;
        font-size: 13px !important;
    }}
    div[data-testid="stHorizontalBlock"]:first-of-type [data-baseweb="select"] svg {{
        fill: {WHITE} !important;
    }}
    div[data-testid="stHorizontalBlock"]:first-of-type button {{
        background: {RED} !important;
        color: {WHITE} !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 700 !important;
    }}

    /* Download button */
    .stDownloadButton > button {{
        background: {BLUE} !important;
        color: {WHITE} !important;
        font-weight: 700 !important;
        border: none !important;
        border-radius: 6px !important;
        padding: 10px 24px !important;
        font-size: 15px !important;
    }}
    .stDownloadButton > button:hover {{
        background: #0f3a61 !important;
    }}
    .section-header {{
        font-size: 11px; font-weight: 700; letter-spacing: 1.5px;
        color: {BLUE}; text-transform: uppercase;
        border-bottom: 2px solid {BLUE}; padding-bottom: 4px;
        margin: 16px 0 10px 0;
    }}
</style>
""", unsafe_allow_html=True)

# ── Load available weeks & markets from DB ────────────────────────────────────
@st.cache_data(ttl=300)
def load_filter_options():
    conn, dialect = get_db_connection()
    weeks_q = """
        SELECT DISTINCT week_ending FROM weekly_store_history
        UNION
        SELECT DISTINCT week_ending FROM weekly_sales
        ORDER BY week_ending
    """
    markets_q = "SELECT DISTINCT co_op FROM stores WHERE co_op IS NOT NULL ORDER BY co_op"
    weeks_df = pd.read_sql(weeks_q, conn)
    markets_df = pd.read_sql(markets_q, conn)
    return sorted(weeks_df['week_ending'].tolist()), sorted(markets_df['co_op'].tolist())

@st.cache_data(ttl=300)
def load_store_list(markets=None):
    conn, dialect = get_db_connection()
    if markets:
        placeholders = ','.join(['?' if 'sqlite' in str(type(conn)) else '%s'] * len(markets))
        q = f"SELECT store_id, co_op FROM stores WHERE co_op IN ({placeholders}) ORDER BY store_id"
        df = pd.read_sql(q, conn, params=markets)
    else:
        df = pd.read_sql("SELECT store_id, co_op FROM stores ORDER BY store_id", conn)
    return df

all_weeks, all_markets = load_filter_options()

# ── Header bar ────────────────────────────────────────────────────────────────
logo_col, from_col, to_col, mkt_col, home_col = st.columns([2.2, 1.5, 1.5, 2.5, 1])

with logo_col:
    st.markdown(f"""
        <div style='display:flex;align-items:center;gap:12px;padding:4px 0;'>
            <img src="{_LOGO}" style="height:38px;width:auto;"/>
            <span style='font-size:13px;font-weight:700;letter-spacing:0.5px;color:white;'>
                DATA EXPORT
            </span>
        </div>
    """, unsafe_allow_html=True)

with from_col:
    from_week = st.selectbox("From Week", all_weeks,
                             index=max(0, len(all_weeks) - 52),
                             key="from_week")

with to_col:
    to_week = st.selectbox("To Week", all_weeks,
                           index=len(all_weeks) - 1,
                           key="to_week")

with mkt_col:
    mkt_options = ["All Markets"] + all_markets
    selected_mkt = st.selectbox("Market", mkt_options, key="export_mkt")

with home_col:
    st.markdown("<div style='padding-top:22px;'></div>", unsafe_allow_html=True)
    st.page_link("app.py", label="⌂ Home", use_container_width=True)

# ── Pull and merge data ───────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_export_data(from_week, to_week, market_filter):
    conn, dialect = get_db_connection()
    p = '%s' if dialect == 'postgres' else '?'

    # 1. weekly_sales (recent, richest columns)
    ws_q = f"""
        SELECT
            ws.week_ending,
            ws.store_id,
            s.co_op       AS market,
            s.city,
            ws.net_sales,
            NULL          AS transactions,
            ws.sss_pct,
            ws.same_store_txn_pct,
            ws.same_store_ticket_pct,
            ws.online_sales_pct,
            ws.third_party_sales_pct,
            ws.loyalty_sales_pct,
            ws.avg_daily_bread,
            ws.fytd_net_sales,
            ws.fytd_sss_pct
        FROM weekly_sales ws
        LEFT JOIN stores s ON ws.store_id = s.store_id
        WHERE ws.week_ending >= {p} AND ws.week_ending <= {p}
    """
    params = [from_week, to_week]
    if market_filter and market_filter != "All Markets":
        ws_q += f" AND s.co_op = {p}"
        params.append(market_filter)

    # 2. weekly_store_history (long history)
    sh_q = f"""
        SELECT
            sh.week_ending,
            sh.store_id,
            s.co_op       AS market,
            s.city,
            sh.net_sales,
            sh.transactions,
            NULL AS sss_pct,
            NULL AS same_store_txn_pct,
            NULL AS same_store_ticket_pct,
            NULL AS online_sales_pct,
            NULL AS third_party_sales_pct,
            NULL AS loyalty_sales_pct,
            NULL AS avg_daily_bread,
            NULL AS fytd_net_sales,
            NULL AS fytd_sss_pct
        FROM weekly_store_history sh
        LEFT JOIN stores s ON sh.store_id = s.store_id
        WHERE sh.week_ending >= {p} AND sh.week_ending <= {p}
    """
    params2 = [from_week, to_week]
    if market_filter and market_filter != "All Markets":
        sh_q += f" AND s.co_op = {p}"
        params2.append(market_filter)

    ws_df = pd.read_sql(ws_q, conn, params=params)
    sh_df = pd.read_sql(sh_q, conn, params=params2)

    # Only keep store_history rows for weeks NOT already in weekly_sales
    ws_weeks = set(ws_df['week_ending'].unique())
    sh_df = sh_df[~sh_df['week_ending'].isin(ws_weeks)]

    combined = pd.concat([ws_df, sh_df], ignore_index=True)

    # 3. Bread supplement (for weeks in weekly_sales that didn't have bread)
    bread_q = f"""
        SELECT wb.week_ending, wb.store_id, wb.bread_count, wb.avg_daily_bread AS bread_avg_daily
        FROM weekly_bread wb
        WHERE wb.week_ending >= {p} AND wb.week_ending <= {p}
    """
    try:
        bread_df = pd.read_sql(bread_q, conn, params=[from_week, to_week])
        combined = combined.merge(bread_df[['week_ending','store_id','bread_count']],
                                  on=['week_ending','store_id'], how='left')
    except Exception:
        combined['bread_count'] = None

    # 4. Loyalty supplement
    loy_q = f"""
        SELECT week_ending, store_id,
               member_activations_current, member_transactions_current,
               points_earned_current, points_redeemed_current
        FROM weekly_loyalty
        WHERE week_ending >= {p} AND week_ending <= {p}
    """
    try:
        loy_df = pd.read_sql(loy_q, conn, params=[from_week, to_week])
        combined = combined.merge(loy_df, on=['week_ending','store_id'], how='left')
    except Exception:
        for c in ['member_activations_current','member_transactions_current',
                  'points_earned_current','points_redeemed_current']:
            combined[c] = None

    # Add store name
    combined['store_name'] = combined['store_id'].map(STORE_NAMES).fillna('')

    # Sort and rename columns for clean Excel output
    combined = combined.sort_values(['week_ending','market','store_id'], na_position='last')
    combined = combined.rename(columns={
        'week_ending':                   'Week Ending',
        'store_id':                      'Store #',
        'store_name':                    'Store Name',
        'city':                          'City',
        'market':                        'Market',
        'net_sales':                     'Net Sales ($)',
        'transactions':                  'Transactions',
        'sss_pct':                       'SSS %',
        'same_store_txn_pct':            'SS Transactions %',
        'same_store_ticket_pct':         'SS Avg Ticket %',
        'online_sales_pct':              'Online Sales %',
        'third_party_sales_pct':         '3rd Party Sales %',
        'loyalty_sales_pct':             'Loyalty Sales %',
        'avg_daily_bread':               'Avg Daily Bread',
        'bread_count':                   'Bread Count',
        'fytd_net_sales':                'FYTD Net Sales ($)',
        'fytd_sss_pct':                  'FYTD SSS %',
        'member_activations_current':    'New Loyalty Members',
        'member_transactions_current':   'Loyalty Transactions',
        'points_earned_current':         'Points Earned',
        'points_redeemed_current':       'Points Redeemed',
    })

    # Reorder columns logically
    col_order = [
        'Week Ending', 'Store #', 'Store Name', 'City', 'Market',
        'Net Sales ($)', 'Transactions',
        'SSS %', 'SS Transactions %', 'SS Avg Ticket %',
        'Online Sales %', '3rd Party Sales %', 'Loyalty Sales %',
        'Avg Daily Bread', 'Bread Count',
        'FYTD Net Sales ($)', 'FYTD SSS %',
        'New Loyalty Members', 'Loyalty Transactions',
        'Points Earned', 'Points Redeemed',
    ]
    col_order = [c for c in col_order if c in combined.columns]
    combined = combined[col_order]

    return combined


# Run query
if from_week > to_week:
    st.error("'From Week' must be on or before 'To Week'.")
    st.stop()

with st.spinner("Loading data…"):
    df = load_export_data(from_week, to_week, selected_mkt)

# ── Summary stats bar ─────────────────────────────────────────────────────────
num_weeks   = df['Week Ending'].nunique()
num_stores  = df['Store #'].nunique()
num_rows    = len(df)

st.markdown(f"""
<div style='display:flex;gap:32px;background:{LIGHT};border:1px solid {BORDER};
            border-radius:8px;padding:12px 20px;margin-bottom:12px;'>
    <div><span style='font-size:22px;font-weight:800;color:{BLUE};'>{num_weeks}</span>
         <span style='font-size:12px;color:#666;margin-left:6px;'>Weeks</span></div>
    <div><span style='font-size:22px;font-weight:800;color:{BLUE};'>{num_stores}</span>
         <span style='font-size:12px;color:#666;margin-left:6px;'>Stores</span></div>
    <div><span style='font-size:22px;font-weight:800;color:{BLUE};'>{num_rows:,}</span>
         <span style='font-size:12px;color:#666;margin-left:6px;'>Rows</span></div>
    <div style='margin-left:auto;font-size:12px;color:#888;align-self:center;'>
        Note: SSS %, ticket %, and other rate metrics are only available for the most recent weeks
        (when weekly sales reports were loaded). Historical weeks show Net Sales + Transactions only.
    </div>
</div>
""", unsafe_allow_html=True)

# ── Build Excel in-memory ─────────────────────────────────────────────────────
def build_excel(data: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers as xlnums)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Store Data"

    NAV_BLUE  = "134A7C"
    JM_RED    = "EE3227"
    JM_GOLD   = "D4AF37"
    HEADER_FG = "FFFFFF"
    ALT_ROW   = "EFF4FA"

    thin = Side(style='thin', color="C0C8D4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ────────────────────────────────────────────────────────────
    headers = list(data.columns)
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, color=HEADER_FG, size=10)
        cell.fill = PatternFill('solid', start_color=NAV_BLUE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    currency_cols = {'Net Sales ($)', 'FYTD Net Sales ($)'}
    pct_cols      = {'SSS %', 'SS Transactions %', 'SS Avg Ticket %',
                     'Online Sales %', '3rd Party Sales %', 'Loyalty Sales %', 'FYTD SSS %'}
    int_cols      = {'Transactions', 'Bread Count', 'New Loyalty Members',
                     'Loyalty Transactions', 'Points Earned', 'Points Redeemed'}

    for ri, row in enumerate(data.itertuples(index=False), start=2):
        fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None
        for ci, (h, val) in enumerate(zip(headers, row), start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='right' if h not in
                                       ('Week Ending','Store #','Store Name','City','Market')
                                       else 'left', vertical='center')
            cell.border = border
            if fill:
                cell.fill = fill
            # Number formats
            if h in currency_cols and val is not None:
                cell.number_format = '$#,##0'
            elif h in pct_cols and val is not None:
                cell.number_format = '+0.0%;-0.0%;"-"'
            elif h in int_cols and val is not None:
                cell.number_format = '#,##0'
            elif h == 'Avg Daily Bread' and val is not None:
                cell.number_format = '#,##0.0'

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        'Week Ending': 13, 'Store #': 10, 'Store Name': 18, 'City': 14, 'Market': 22,
        'Net Sales ($)': 14, 'Transactions': 14, 'SSS %': 10, 'SS Transactions %': 16,
        'SS Avg Ticket %': 16, 'Online Sales %': 14, '3rd Party Sales %': 16,
        'Loyalty Sales %': 14, 'Avg Daily Bread': 15, 'Bread Count': 13,
        'FYTD Net Sales ($)': 18, 'FYTD SSS %': 12, 'New Loyalty Members': 18,
        'Loyalty Transactions': 20, 'Points Earned': 15, 'Points Redeemed': 16,
    }
    for ci, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 14)

    # ── Freeze top row ────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Cover / metadata tab ─────────────────────────────────────────────────
    info = wb.create_sheet("Export Info")
    info_data = [
        ["Jersey Mike's Valley Group — Weekly Store Data Export"],
        [],
        ["Date Range",    f"{from_week}  →  {to_week}"],
        ["Market Filter", selected_mkt],
        ["Weeks",         num_weeks],
        ["Stores",        num_stores],
        ["Total Rows",    num_rows],
        [],
        ["Column Notes"],
        ["Net Sales ($)",       "Weekly net sales in dollars"],
        ["Transactions",        "Weekly transaction count (historical data only, prior to recent reports)"],
        ["SSS %",               "Same Store Sales % vs prior year (recent report weeks only)"],
        ["SS Transactions %",   "Same Store Transactions % vs prior year (recent report weeks only)"],
        ["SS Avg Ticket %",     "Same Store Avg Ticket % vs prior year (recent report weeks only)"],
        ["Online Sales %",      "Online sales as % of net sales"],
        ["3rd Party Sales %",   "3rd party delivery as % of net sales"],
        ["Loyalty Sales %",     "Loyalty member sales as % of net sales"],
        ["FYTD Net Sales ($)",  "Fiscal year-to-date net sales"],
        ["FYTD SSS %",          "Fiscal year-to-date same store sales %"],
    ]
    for r, row_data in enumerate(info_data, start=1):
        for c, val in enumerate(row_data, start=1):
            cell = info.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(name='Arial', bold=True, size=13, color=NAV_BLUE)
            elif r == 9:
                cell.font = Font(name='Arial', bold=True, size=11, color=NAV_BLUE)
            elif c == 1 and r > 2:
                cell.font = Font(name='Arial', bold=True, size=10)
            else:
                cell.font = Font(name='Arial', size=10)
    info.column_dimensions['A'].width = 24
    info.column_dimensions['B'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Download button ───────────────────────────────────────────────────────────
st.markdown('<div class="section-header">EXPORT</div>', unsafe_allow_html=True)

dl_col, info_col = st.columns([2, 5])
with dl_col:
    fname = f"JMV_Weekly_Store_Data_{from_week}_to_{to_week}.xlsx"
    excel_bytes = build_excel(df)
    st.download_button(
        label=f"⬇ Download Excel ({num_rows:,} rows)",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

with info_col:
    st.markdown(f"""
    <div style='padding:10px 0;font-size:13px;color:#555;'>
        <b>File:</b> {fname}<br>
        Includes an <b>Export Info</b> tab with column descriptions.
        All columns with <b>auto-filter</b> and frozen header row for easy slicing in Excel.
    </div>
    """, unsafe_allow_html=True)

# ── Data preview ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">DATA PREVIEW (first 200 rows)</div>', unsafe_allow_html=True)

preview = df.head(200).copy()

# Format for display
for col in ['Net Sales ($)', 'FYTD Net Sales ($)']:
    if col in preview.columns:
        preview[col] = preview[col].apply(lambda x: f"${x:,.0f}" if pd.notna(x) else "—")

for col in ['SSS %', 'SS Transactions %', 'SS Avg Ticket %',
            'Online Sales %', '3rd Party Sales %', 'Loyalty Sales %', 'FYTD SSS %']:
    if col in preview.columns:
        preview[col] = preview[col].apply(
            lambda x: (f"+{x:.1f}%" if x >= 0 else f"{x:.1f}%") if pd.notna(x) else "—"
        )

for col in ['Transactions', 'Bread Count', 'New Loyalty Members',
            'Loyalty Transactions', 'Points Earned', 'Points Redeemed']:
    if col in preview.columns:
        preview[col] = preview[col].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")

st.dataframe(preview, use_container_width=True, height=480, hide_index=True)
